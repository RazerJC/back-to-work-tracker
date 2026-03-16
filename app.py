"""
Employee Attendance Analytics System — Flask Backend
"""
import os
import io
import re
import json
import tempfile
from datetime import datetime, date
from collections import defaultdict

from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB

# ── In-memory data store ──────────────────────────────────────────────
DATA_STORE = {
    "records": [],          # list of dicts
    "loaded": False,
    "source_file": None,
}

# ── Helpers ───────────────────────────────────────────────────────────

HEADER_MAP = {
    "NAME": "name",
    "DATE OF ABSENCE": "date_of_absence",
    "DATE OF ABSENCE\n(MONTH/DAY/YEAR)": "date_of_absence_alt",
    "NUMBER OF WORKING DAY/(S) ABSENT": "days_absent",
    "NUMBER OF WORKING DAYS ABSENT": "days_absent",
    "JOB TITLE / POSITION :": "position",
    "JOB TITLE / POSITION": "position",
    "DEPARTMENT :": "department",
    "DEPARTMENT": "department",
    "REASON :": "reason",
    "REASON": "reason",
    "DATE OF INTENDED RETURN TO WORK :": "return_date",
    "TIMESTAMP": "timestamp",
    "COMPANY NAME": "company_name",
}

# Headers to explicitly ignore (should NOT be mapped to any field)
IGNORED_HEADERS = {
    "EMAIL ADDRESS", "PHONE NUMBER", "HR NAME",
    "DATE TODAY", "ALLOW EMPLOYEE TO RTW",
}


def _normalize_header(h):
    if not h:
        return None
    h_clean = h.strip().upper().replace("\n", "\n")
    # Skip ignored headers first
    for ignored in IGNORED_HEADERS:
        if ignored in h_clean:
            return None
    # Sort patterns by length descending so more specific patterns match first
    sorted_patterns = sorted(HEADER_MAP.items(), key=lambda x: len(x[0]), reverse=True)
    # First try exact match
    for pattern, key in sorted_patterns:
        if pattern.upper() == h_clean:
            return key
    # Then try substring: only check if pattern is IN header (not reverse)
    for pattern, key in sorted_patterns:
        if pattern.upper() in h_clean:
            return key
    return None


def _parse_date(val):
    """Return a date object from various formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    # Try direct format parsing
    for fmt in ("%m/%d/%Y", "%B %d, %Y", "%b %d, %Y", "%Y-%m-%d", "%m/%d/%y",
                 "%B %d,%Y", "%b %d,%Y", "%m-%d-%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Try parsing comma-separated dates (take the first one)
    if ',' in s:
        parts = s.split(',')
        for part in parts:
            result = _parse_date(part.strip())
            if result:
                return result
    return None


def _parse_days(val):
    if val is None:
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        pass
    # Try to extract a number from text like "1 DAY", "4 days", "22 DAYS"
    s = str(val).strip()
    m = re.search(r'(\d+\.?\d*)', s)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    return 0


def _status(total_absences):
    if total_absences >= 5:
        return "Critical"
    if total_absences >= 3:
        return "Warning"
    return "Normal"


# Known company/org patterns to filter out
_COMPANY_PATTERNS = [
    "multimix", "manufacturing", "corp", "inc.", "corporation",
    "company", "llc", "ltd", "enterprises", "sumiden", "circuits",
    "cainiao", "bi chain",
]


def _is_company_name(name):
    """Return True if the name looks like a company/org name."""
    lower = name.lower()
    return any(p in lower for p in _COMPANY_PATTERNS)


def _normalize_name(name):
    """Normalize employee name — handle 'Last, First' format, strip middle initials, and title-case."""
    if not name:
        return None
    cleaned = " ".join(str(name).strip().split())
    if _is_company_name(cleaned):
        return None

    # Handle "LASTNAME, FIRSTNAME" or "LASTNAME, FIRSTNAME M." format
    # Split on first comma only
    if ',' in cleaned:
        parts = cleaned.split(',', 1)
        last = parts[0].strip()
        first = parts[1].strip()
        if first and last:
            cleaned = f"{first} {last}"

    # Remove middle initials: single letter optionally followed by period
    # e.g., "Melodia A. Grenio" -> "Melodia Grenio"
    cleaned = re.sub(r'\b[A-Za-z]\.\s*', '', cleaned)
    cleaned = " ".join(cleaned.split())  # clean up extra spaces

    result = cleaned.title()

    # Handle reversed names without commas (e.g., "Grenio Melodia" vs "Melodia Grenio")
    # We canonicalize by sorting the name parts so "Grenio Melodia" == "Melodia Grenio" 
    # Only do this for exactly 2-part names
    parts = result.split()
    if len(parts) == 2:
        # Always store as sorted pair to deduplicate
        canonical = " ".join(sorted(parts))
        return canonical

    return result


def _normalize_company(val):
    """Normalize company name to a consistent format."""
    if not val:
        return "Unknown"
    s = " ".join(str(val).strip().split()).upper()
    # Normalize known company names
    if "MULTIMIX" in s:
        return "Multimix International Manufacturing Corp."
    if "SUMIDEN" in s:
        return "First Sumiden Circuits Inc."
    if "BI CHAIN" in s:
        return "BI Chain Inc. Philippine"
    if "CAINIAO" in s:
        return "Cainiao"
    return " ".join(str(val).strip().split()).title()


def _normalize_dept(dept):
    if not dept:
        return "Unknown"
    d = " ".join(str(dept).strip().split()).upper()
    # Consolidate common typos
    mapping = {
        "BUNDLIING": "BUNDLING", "BUNDLES": "BUNDLING", "BUNDLNG": "BUNDLING",
        "BUNDLER": "BUNDLING", "BONDLING": "BUNDLING",
        "BUNDLING / VAS": "BUNDLING", "BUNDLING VAS": "BUNDLING",
        "BUNDLING DEPT.": "BUNDLING",
        "HDA COIL - OLD LINE": "HDA COIL", "HDA COIL OLD LINE": "HDA COIL",
        "HDA COIL OLDLINE": "HDA COIL", "HDA-COIL": "HDA COIL",
        "HDA COIIL": "HDA COIL", "HDA": "HDA COIL",
        "COI": "HDA COIL",
    }
    return mapping.get(d, d).title()


def _get_month_days(month, year):
    """Get the number of days in a given month/year."""
    import calendar
    if not month or not year:
        return 31
    return calendar.monthrange(year, month)[1]


def process_workbook(wb):
    """Process an openpyxl Workbook and return a list of record dicts."""
    ws = wb.active
    records = []

    # Map column indices to field names
    col_map = {}
    for ci, cell in enumerate(ws[1], 1):
        key = _normalize_header(str(cell.value) if cell.value else "")
        if key:
            col_map[ci] = key

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = {col_map.get(ci): cell.value for ci, cell in enumerate(row, 1) if ci in col_map}

        name = _normalize_name(vals.get("name"))
        if not name:
            continue

        days = _parse_days(vals.get("days_absent"))
        if days <= 0:
            continue

        # Try multiple date sources
        absence_date = _parse_date(vals.get("date_of_absence"))
        if not absence_date:
            absence_date = _parse_date(vals.get("date_of_absence_alt"))
        # Fallback: use the timestamp (submission date) if available
        if not absence_date and vals.get("timestamp"):
            absence_date = _parse_date(vals.get("timestamp"))

        # Cap days_absent to the number of working days in the month (max 23)
        if days > 23:
            days = 23  # Max working days in any month

        dept = _normalize_dept(vals.get("department"))
        position = str(vals.get("position", "")).strip().title() if vals.get("position") else "N/A"
        reason = str(vals.get("reason", "")).strip() if vals.get("reason") else ""
        company = _normalize_company(vals.get("company_name"))

        month = absence_date.month if absence_date else None
        year = absence_date.year if absence_date else None
        month_name = absence_date.strftime("%B") if absence_date else "Unknown"

        records.append({
            "name": name,
            "company": company,
            "department": dept,
            "position": position,
            "date_of_absence": absence_date.isoformat() if absence_date else None,
            "month": month,
            "month_name": month_name,
            "year": year,
            "days_absent": days,
            "reason": reason,
        })

    return records


# ── Routes ────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/upload", methods=["POST"])
def upload():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    try:
        wb = load_workbook(io.BytesIO(f.read()), data_only=True)
        records = process_workbook(wb)
        wb.close()
        DATA_STORE["records"] = records
        DATA_STORE["loaded"] = True
        DATA_STORE["source_file"] = f.filename
        return jsonify({"message": f"Processed {len(records)} records", "count": len(records)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/auto_load", methods=["POST"])
def auto_load():
    """Auto-load the default Excel file from the same directory."""
    base = os.path.dirname(os.path.abspath(__file__))
    candidates = [f for f in os.listdir(base) if f.lower().endswith(".xlsx") and not f.startswith("~")]
    if not candidates:
        return jsonify({"error": "No .xlsx file found in application directory"}), 404
    filepath = os.path.join(base, candidates[0])
    try:
        wb = load_workbook(filepath, data_only=True)
        records = process_workbook(wb)
        wb.close()
        DATA_STORE["records"] = records
        DATA_STORE["loaded"] = True
        DATA_STORE["source_file"] = candidates[0]
        return jsonify({"message": f"Loaded {len(records)} records from {candidates[0]}", "count": len(records), "filename": candidates[0]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/data")
def get_data():
    if not DATA_STORE["loaded"]:
        return jsonify({"records": [], "loaded": False})

    records = DATA_STORE["records"]
    # Apply filters
    search = request.args.get("search", "").strip().lower()
    month = request.args.get("month", "")
    year = request.args.get("year", "")
    dept = request.args.get("department", "")
    company = request.args.get("company", "")

    if search:
        records = [r for r in records if search in r["name"].lower()]
    if month:
        try:
            records = [r for r in records if r["month"] == int(month)]
        except ValueError:
            pass
    if year:
        try:
            records = [r for r in records if r["year"] == int(year)]
        except ValueError:
            pass
    if dept:
        records = [r for r in records if r["department"].lower() == dept.lower()]
    if company:
        records = [r for r in records if r["company"].lower() == company.lower()]

    # Aggregate by employee + month + year
    agg = defaultdict(lambda: {"days": 0, "dates": [], "reasons": [], "department": "", "position": "", "company": ""})
    for r in records:
        key = (r["name"], r["month_name"], r["year"])
        agg[key]["days"] += r["days_absent"]
        agg[key]["department"] = r["department"]
        agg[key]["position"] = r["position"]
        agg[key]["company"] = r["company"]
        if r["date_of_absence"]:
            agg[key]["dates"].append(r["date_of_absence"])
        if r.get("reason") and r["reason"] not in agg[key]["reasons"]:
            agg[key]["reasons"].append(r["reason"])

    table = []
    for (name, month_name, year_val), v in agg.items():
        total = v["days"]
        # Cap total to max working days in a month (23)
        total = min(total, 23)
        table.append({
            "name": name,
            "company": v["company"],
            "department": v["department"],
            "position": v["position"],
            "month": month_name,
            "year": year_val,
            "total_absences": total,
            "reason": "; ".join(v["reasons"]) if v["reasons"] else "N/A",
            "status": _status(total),
            "dates": sorted(v["dates"]),
        })

    table.sort(key=lambda x: (-x["total_absences"], x["name"]))
    return jsonify({"records": table, "loaded": True})


@app.route("/api/stats")
def get_stats():
    if not DATA_STORE["loaded"]:
        return jsonify({})

    records = DATA_STORE["records"]
    search = request.args.get("search", "").strip().lower()
    month = request.args.get("month", "")
    year = request.args.get("year", "")
    dept = request.args.get("department", "")
    company = request.args.get("company", "")

    if search:
        records = [r for r in records if search in r["name"].lower()]
    if month:
        try:
            records = [r for r in records if r["month"] == int(month)]
        except ValueError:
            pass
    if year:
        try:
            records = [r for r in records if r["year"] == int(year)]
        except ValueError:
            pass
    if dept:
        records = [r for r in records if r["department"].lower() == dept.lower()]
    if company:
        records = [r for r in records if r["company"].lower() == company.lower()]

    total_records = len(records)
    total_days = sum(r["days_absent"] for r in records)
    unique_employees = len(set(r["name"] for r in records))

    # Aggregate per employee per month for status counting
    emp_month = defaultdict(float)
    for r in records:
        emp_month[(r["name"], r["month"], r["year"])] += r["days_absent"]

    # Cap each employee-month to 23
    for key in emp_month:
        emp_month[key] = min(emp_month[key], 23)

    warnings = sum(1 for v in emp_month.values() if 3 <= v < 5)
    criticals = sum(1 for v in emp_month.values() if v >= 5)

    # Department breakdown
    dept_totals = defaultdict(float)
    for r in records:
        dept_totals[r["department"]] += r["days_absent"]
    dept_chart = sorted(dept_totals.items(), key=lambda x: -x[1])

    # Monthly trends
    month_totals = defaultdict(float)
    for r in records:
        if r["month"] and r["year"]:
            label = f"{r['year']}-{str(r['month']).zfill(2)}"
            month_totals[label] += r["days_absent"]
    month_chart = sorted(month_totals.items())

    # Top 10 employees
    emp_totals = defaultdict(float)
    for r in records:
        emp_totals[r["name"]] += r["days_absent"]
    top10 = sorted(emp_totals.items(), key=lambda x: -x[1])[:10]

    # Company breakdown
    company_totals = defaultdict(lambda: {"days": 0, "employees": set(), "departments": set()})
    for r in records:
        company_totals[r["company"]]["days"] += r["days_absent"]
        company_totals[r["company"]]["employees"].add(r["name"])
        company_totals[r["company"]]["departments"].add(r["department"])

    company_data = []
    for cname, cdata in sorted(company_totals.items(), key=lambda x: -x[1]["days"]):
        company_data.append({
            "name": cname,
            "total_days": cdata["days"],
            "employee_count": len(cdata["employees"]),
            "department_count": len(cdata["departments"]),
            "departments": sorted(cdata["departments"]),
        })

    # Available filters
    all_records = DATA_STORE["records"]
    months = sorted(set(r["month"] for r in all_records if r["month"]))
    years = sorted(set(r["year"] for r in all_records if r["year"]))
    depts = sorted(set(r["department"] for r in all_records))
    companies = sorted(set(r["company"] for r in all_records))

    return jsonify({
        "total_records": total_records,
        "total_days": total_days,
        "unique_employees": unique_employees,
        "warnings": warnings,
        "criticals": criticals,
        "dept_chart": {"labels": [d[0] for d in dept_chart], "values": [d[1] for d in dept_chart]},
        "month_chart": {"labels": [m[0] for m in month_chart], "values": [m[1] for m in month_chart]},
        "top10": {"labels": [t[0] for t in top10], "values": [t[1] for t in top10]},
        "companies": company_data,
        "filters": {"months": months, "years": years, "departments": depts, "companies": companies},
    })

@app.route("/api/employees")
def get_employees():
    """Get all unique employees with their totals."""
    if not DATA_STORE["loaded"]:
        return jsonify({"employees": []})

    records = DATA_STORE["records"]
    search = request.args.get("search", "").strip().lower()
    month = request.args.get("month", "")
    year = request.args.get("year", "")
    dept = request.args.get("department", "")
    company = request.args.get("company", "")

    if search:
        records = [r for r in records if search in r["name"].lower()]
    if month:
        try:
            records = [r for r in records if r["month"] == int(month)]
        except ValueError:
            pass
    if year:
        try:
            records = [r for r in records if r["year"] == int(year)]
        except ValueError:
            pass
    if dept:
        records = [r for r in records if r["department"].lower() == dept.lower()]
    if company:
        records = [r for r in records if r["company"].lower() == company.lower()]

    emp_data = defaultdict(lambda: {"days": 0, "company": "", "department": "", "position": "", "reasons": []})
    for r in records:
        emp_data[r["name"]]["days"] += r["days_absent"]
        emp_data[r["name"]]["company"] = r["company"]
        emp_data[r["name"]]["department"] = r["department"]
        emp_data[r["name"]]["position"] = r["position"]
        if r.get("reason") and r["reason"] not in emp_data[r["name"]]["reasons"]:
            emp_data[r["name"]]["reasons"].append(r["reason"])

    employees = []
    for ename, edata in sorted(emp_data.items(), key=lambda x: -x[1]["days"]):
        total = min(edata["days"], 23)
        employees.append({
            "name": ename,
            "company": edata["company"],
            "department": edata["department"],
            "position": edata["position"],
            "total_days": total,
            "reason": "; ".join(edata["reasons"]) if edata["reasons"] else "N/A",
            "status": _status(total),
        })

    return jsonify({"employees": employees, "count": len(employees)})


@app.route("/api/company/<name>")
def company_detail(name):
    """Get detailed breakdown for a specific company."""
    if not DATA_STORE["loaded"]:
        return jsonify({"error": "No data loaded"}), 400

    records = [r for r in DATA_STORE["records"] if r["company"].lower() == name.lower()]
    if not records:
        return jsonify({"error": "Company not found"}), 404

    # Department breakdown within company
    dept_data = defaultdict(lambda: {"days": 0, "employees": set(), "records": 0})
    for r in records:
        dept_data[r["department"]]["days"] += r["days_absent"]
        dept_data[r["department"]]["employees"].add(r["name"])
        dept_data[r["department"]]["records"] += 1

    departments = []
    for dname, ddata in sorted(dept_data.items(), key=lambda x: -x[1]["days"]):
        departments.append({
            "name": dname,
            "total_days": ddata["days"],
            "employee_count": len(ddata["employees"]),
            "record_count": ddata["records"],
        })

    total_days = sum(r["days_absent"] for r in records)
    unique_employees = len(set(r["name"] for r in records))

    return jsonify({
        "company": name,
        "total_days": total_days,
        "unique_employees": unique_employees,
        "total_records": len(records),
        "departments": departments,
    })


@app.route("/api/department/<company>/<dept>")
def department_detail(company, dept):
    """Get employee breakdown for a specific department within a company."""
    if not DATA_STORE["loaded"]:
        return jsonify({"error": "No data loaded"}), 400

    records = [r for r in DATA_STORE["records"]
               if r["company"].lower() == company.lower() and r["department"].lower() == dept.lower()]
    if not records:
        return jsonify({"error": "Department not found"}), 404

    # Employee breakdown
    emp_data = defaultdict(lambda: {"days": 0, "position": "", "records": 0, "dates": [], "reasons": []})
    for r in records:
        emp_data[r["name"]]["days"] += r["days_absent"]
        emp_data[r["name"]]["position"] = r["position"]
        emp_data[r["name"]]["records"] += 1
        if r["date_of_absence"]:
            emp_data[r["name"]]["dates"].append(r["date_of_absence"])
        if r.get("reason") and r["reason"] not in emp_data[r["name"]]["reasons"]:
            emp_data[r["name"]]["reasons"].append(r["reason"])

    employees = []
    for ename, edata in sorted(emp_data.items(), key=lambda x: -x[1]["days"]):
        total = min(edata["days"], 23 * edata["records"])  # reasonable cap
        employees.append({
            "name": ename,
            "position": edata["position"],
            "total_days": edata["days"],
            "record_count": edata["records"],
            "reason": "; ".join(edata["reasons"]) if edata["reasons"] else "N/A",
            "status": _status(edata["days"]),
            "dates": sorted(edata["dates"]),
        })

    return jsonify({
        "company": company,
        "department": dept,
        "total_days": sum(r["days_absent"] for r in records),
        "unique_employees": len(emp_data),
        "employees": employees,
    })


@app.route("/api/employee/<name>")
def employee_profile(name):
    if not DATA_STORE["loaded"]:
        return jsonify({"error": "No data loaded"}), 400

    records = [r for r in DATA_STORE["records"] if r["name"].lower() == name.lower()]
    if not records:
        return jsonify({"error": "Employee not found"}), 404

    emp = records[0]
    total_days = sum(r["days_absent"] for r in records)
    all_dates = sorted(set(r["date_of_absence"] for r in records if r["date_of_absence"]))
    all_reasons = list(set(r["reason"] for r in records if r.get("reason")))

    # Monthly breakdown with individual date+reason pairs
    month_data = defaultdict(lambda: {"days": 0.0, "reasons": [], "date_details": []})
    for r in records:
        if r["month_name"] and r["year"]:
            key = f"{r['month_name']} {r['year']}"
            month_data[key]["days"] += r["days_absent"]
            if r.get("reason") and r["reason"] not in month_data[key]["reasons"]:
                month_data[key]["reasons"].append(r["reason"])
            # Store individual date + reason pair
            if r["date_of_absence"]:
                month_data[key]["date_details"].append({
                    "date": r["date_of_absence"],
                    "day": int(r["date_of_absence"].split("-")[2]) if r["date_of_absence"] else None,
                    "days_absent": r["days_absent"],
                    "reason": r.get("reason", "") or "N/A",
                })

    # Cap each month to 23
    for key in month_data:
        month_data[key]["days"] = min(month_data[key]["days"], 23)

    # Status per month
    monthly_list = []
    for label, mdata in sorted(month_data.items()):
        # Sort date_details by date
        sorted_dates = sorted(mdata["date_details"], key=lambda x: x["date"])
        monthly_list.append({
            "period": label,
            "days": mdata["days"],
            "reason": "; ".join(mdata["reasons"]) if mdata["reasons"] else "N/A",
            "status": _status(mdata["days"]),
            "dates": sorted_dates,
        })

    # Recalculate total from capped values
    total_days = sum(m["days"] for m in monthly_list)

    return jsonify({
        "name": emp["name"],
        "company": emp["company"],
        "department": emp["department"],
        "position": emp["position"],
        "total_days": total_days,
        "total_records": len(records),
        "absence_dates": all_dates,
        "reasons": all_reasons,
        "monthly_breakdown": monthly_list,
        "overall_status": _status(total_days),
        "chart": {"labels": [m["period"] for m in monthly_list], "values": [m["days"] for m in monthly_list]},
    })


@app.route("/api/alerts")
def get_alerts():
    if not DATA_STORE["loaded"]:
        return jsonify({"alerts": []})

    emp_month = defaultdict(lambda: {"days": 0, "dept": "", "company": "", "month": "", "year": None})
    for r in DATA_STORE["records"]:
        key = (r["name"], r["month_name"], r["year"])
        emp_month[key]["days"] += r["days_absent"]
        emp_month[key]["dept"] = r["department"]
        emp_month[key]["company"] = r["company"]
        emp_month[key]["month"] = r["month_name"]
        emp_month[key]["year"] = r["year"]

    alerts = []
    for (name, month_name, year), v in emp_month.items():
        days = min(v["days"], 23)  # cap to max working days
        if days >= 3:
            alerts.append({
                "name": name,
                "company": v["company"],
                "department": v["dept"],
                "month": f"{month_name} {year}" if year else month_name,
                "total_absences": days,
                "status": _status(days),
            })

    alerts.sort(key=lambda x: (-x["total_absences"], x["name"]))
    return jsonify({"alerts": alerts})


@app.route("/api/export/<report_type>")
def export_report(report_type):
    if not DATA_STORE["loaded"]:
        return jsonify({"error": "No data loaded"}), 400

    wb = Workbook()
    ws = wb.active

    if report_type == "monthly":
        ws.title = "Monthly Absence Report"
        ws.append(["Employee Name", "Company", "Department", "Month", "Year", "Total Absences", "Status"])
        agg = defaultdict(lambda: {"days": 0, "dept": "", "company": ""})
        for r in DATA_STORE["records"]:
            key = (r["name"], r["month_name"], r["year"])
            agg[key]["days"] += r["days_absent"]
            agg[key]["dept"] = r["department"]
            agg[key]["company"] = r["company"]
        for (name, month_name, year), v in sorted(agg.items()):
            days = min(v["days"], 23)
            ws.append([name, v["company"], v["dept"], month_name, year, days, _status(days)])

    elif report_type == "department":
        ws.title = "Department Report"
        ws.append(["Company", "Department", "Total Absences", "Unique Employees"])
        dept_data = defaultdict(lambda: {"days": 0, "emps": set()})
        for r in DATA_STORE["records"]:
            dept_data[(r["company"], r["department"])]["days"] += r["days_absent"]
            dept_data[(r["company"], r["department"])]["emps"].add(r["name"])
        for (company, dept), v in sorted(dept_data.items(), key=lambda x: -x[1]["days"]):
            ws.append([company, dept, v["days"], len(v["emps"])])

    elif report_type == "leaderboard":
        ws.title = "Absence Leaderboard"
        ws.append(["Rank", "Employee Name", "Company", "Department", "Total Absences", "Status"])
        emp_data = defaultdict(lambda: {"days": 0, "dept": "", "company": ""})
        for r in DATA_STORE["records"]:
            emp_data[r["name"]]["days"] += r["days_absent"]
            emp_data[r["name"]]["dept"] = r["department"]
            emp_data[r["name"]]["company"] = r["company"]
        for rank, (name, v) in enumerate(sorted(emp_data.items(), key=lambda x: -x[1]["days"]), 1):
            ws.append([rank, name, v["company"], v["dept"], v["days"], _status(v["days"])])

    elif report_type == "critical":
        ws.title = "Critical Absence Report"
        ws.append(["Employee Name", "Company", "Department", "Month", "Year", "Total Absences", "Status", "Warning Message"])
        agg = defaultdict(lambda: {"days": 0, "dept": "", "company": ""})
        for r in DATA_STORE["records"]:
            key = (r["name"], r["month_name"], r["year"])
            agg[key]["days"] += r["days_absent"]
            agg[key]["dept"] = r["department"]
            agg[key]["company"] = r["company"]
        for (name, month_name, year), v in sorted(agg.items()):
            days = min(v["days"], 23)
            status = _status(days)
            if status in ("Warning", "Critical"):
                msg = f"Employee {name} from {v['company']} / {v['dept']} has {days} absence(s) in {month_name} {year}. Status: {status} Absenteeism."
                ws.append([name, v["company"], v["dept"], month_name, year, days, status, msg])

    elif report_type == "email_warnings":
        ws.title = "Email Warning Reports"
        ws.append(["To", "Subject", "Body"])
        agg = defaultdict(lambda: {"days": 0, "dept": "", "company": ""})
        for r in DATA_STORE["records"]:
            key = (r["name"], r["month_name"], r["year"])
            agg[key]["days"] += r["days_absent"]
            agg[key]["dept"] = r["department"]
            agg[key]["company"] = r["company"]
        for (name, month_name, year), v in sorted(agg.items()):
            days = min(v["days"], 23)
            status = _status(days)
            if status in ("Warning", "Critical"):
                subject = f"Employee Absence Warning - {name}"
                body = (
                    f"Employee: {name}\n"
                    f"Company: {v['company']}\n"
                    f"Department: {v['dept']}\n"
                    f"Month: {month_name} {year}\n"
                    f"Total Absences: {days}\n"
                    f"Status: {status} Absenteeism\n\n"
                    f"This is an automated warning report generated by the Employee Attendance Analytics System."
                )
                ws.append(["HR Department", subject, body])
    else:
        return jsonify({"error": "Invalid report type"}), 400

    # Style header
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"HR_{report_type}_report.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Run ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=True, port=5000)
